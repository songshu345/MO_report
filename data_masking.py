# coding = utf-8

import pandas as pd
import os
import openpyxl
import xlsxwriter
import numpy as np

# 文件夹路径
excel_work = 'D:/treasture/cloud_ama/report_automation/mo_masking/'  # Excel文件所在文件夹
masking_path = 'D:/treasture/cloud_ama/report_automation/masking.xlsx'  # 解密文件路径
# Excel转为Pdf
reporters = []


# 报告人员列表
def reporters_list():
    for excel_file in os.listdir(excel_work):
        excel_path = excel_work + excel_file  # excel文件路径
        filePath = os.path.join(excel_path)
        wb = openpyxl.load_workbook(filePath)  # 打开excel文件(Excel整个工作簿)
        sheet_name = wb.sheetnames[0]  # sheet
        ws = wb[sheet_name]
        res = ws.cell(row=39, column=4).value  # 获取单元格的内容
        reporter = res.split('：')[1]
        reporters.append(reporter)  # 汇报人列表
    return reporters


#
reporters = reporters_list()
print(reporters)


# 报告人员解密(解密部分，图片部门未被提取重写)
def masking_reporter(masking_path):
    masking_file = pd.read_excel(masking_path, header=0, index_col=None)  # sheet参数为None，返回的是全表
    L = len(reporters)
    for i in range(L):
        re = reporters[i-1].strip(" ")
        print(re)
        index = masking_file['index']
        name = masking_file['name']
        # 索引所在行
        row = masking_file[masking_file.index == i].index.tolist()
        real_name_row = name[row]
        print(real_name_row)
        real_name = real_name_row.values
        new_report = real_name
        new_report = np.array(new_report)
        new_reporter = new_report.tolist()
        rep_masking = '被评估人： ' + str(new_reporter[0])
        print(rep_masking)
        # 打开指定Excel文件中
        wb = openpyxl.load_workbook(excel_work+re+'.xlsx')  # 打开excel文件
        sheet_name = wb.sheetnames[0]  # sheet
        ws = wb[sheet_name]
        ws.cell(39,4,rep_masking)
        wb.save(excel_work+re+'.xlsx')

    return masking_file


masking_file = masking_reporter(masking_path)
# print(masking_file)
